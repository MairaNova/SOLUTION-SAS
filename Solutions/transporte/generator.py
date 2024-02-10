import datetime
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Border, PatternFill, Font, Alignment

class XlsGenerator:

    def __init__(self, author=None, sheet_name=None,
                 name_app='', color_header='365F91',
                 color_text_header='FFFFFF',
                 header_alignment_h='center',
                 header_alignment_v='center', tabColor='00ff7f',
                 font_type='Arial'):
        '''
        :param author: str
        :param sheet_name: list()
        :param filename: str
        :param name_app: str
        '''
        if not sheet_name:
            raise ValueError('Empty sheet_name')
        if not author:
            raise ValueError('Empty author')
        self.__sheet_name = sheet_name
        now = datetime.datetime.now()
        self.__author = f'{author} - {now.year}/{now.month}/{now.day} ' \
                        f'{now.hour}:{now.minute}'
        self.__name_app = name_app
        self.__row = 1
        self.__fill_header = PatternFill(start_color=color_header,
                                         end_color=color_header,
                                         fill_type="solid")
        self.__color_text_header = Font(color=color_text_header, name=font_type)
        self.__header_alignment = Alignment(horizontal=header_alignment_h,
                                            vertical=header_alignment_v)
        self.__font_type = font_type
        self.__work_book = Workbook()
        self.current_sheet = self.__work_book.active
        self.current_sheet.sheet_properties.tabColor = tabColor
        self.current_sheet.title = self.__sheet_name[0]
        self.sheets = self.__create_sheet()

    def __create_sheet(self):
        yield self.current_sheet
        for sheet in range(1, len(self.__sheet_name)):
            yield self.__work_book.create_sheet(title=self.__sheet_name[sheet])

    def __header(self, title, fields_length=None, fields=None):
        '''
        :param title:
        :param fields_length:
        :return:
        '''
        if not fields_length and fields:
            fields_length = self.__length_fields_title(fields)
        if not title:
            raise ValueError('Empty title')
        self.last_col = get_column_letter(fields_length)
        self.current_sheet[f'A{self.__row}'].fill = self.__fill_header
        self.current_sheet[f'A{self.__row}'] = self.__name_app.upper()
        self.current_sheet[f'A{self.__row}'].font = self.__color_text_header
        self.current_sheet[f'A{self.__row}'].alignment = self.__header_alignment
        self.current_sheet.merge_cells(f'A{self.__row}:{self.last_col}{self.__row+1}')
        self.__row += 3
        self.current_sheet[f'A{self.__row}'] = title.upper()
        self.current_sheet[f'A{self.__row}'].fill = self.__fill_header
        self.current_sheet[f'A{self.__row}'].font = self.__color_text_header
        self.current_sheet[f'A{self.__row}'].alignment = self.__header_alignment
        self.current_sheet.merge_cells(f'A{self.__row}:{self.last_col}{self.__row+1}')
        self.__row += 3
        self.current_sheet[f'A{self.__row}'] = 'Descargado por:'
        self.current_sheet[f'A{self.__row}'].fill = self.__fill_header
        self.current_sheet[f'A{self.__row}'].font = self.__color_text_header
        self.current_sheet[
            f'A{self.__row}'].alignment = self.__header_alignment
        self.current_sheet.merge_cells(f'A{self.__row}:B{self.__row}')
        self.current_sheet[f'C{self.__row}'] = self.__author.upper()
        self.current_sheet[f'C{self.__row}'].fill = self.__fill_header
        self.current_sheet[f'C{self.__row}'].font = self.__color_text_header
        self.current_sheet[
            f'C{self.__row}'].alignment = Alignment(horizontal='left',
                                            vertical='center')
        self.current_sheet.merge_cells(
            f'C{self.__row}:{self.last_col}{self.__row}')
        self.__row += 2

    def __fields_name(self, fields, fill_fields, fond_fields,
                      fields_alignment_h, fields_alignment_v):
        '''
        :param fields:
        :param fields_length:
        :return:
        '''
        if not fields:
            raise ValueError('Empty fields')
        cell = 1
        fill = PatternFill(start_color=fill_fields,
                           end_color=fill_fields,
                           fill_type="solid")
        font = Font(color=fond_fields, name=self.__font_type)
        aligment = Alignment(horizontal=fields_alignment_h,
                             vertical=fields_alignment_v)
        for field in fields:
            column = get_column_letter(cell)
            width = len(str(field)) * 4
            self.current_sheet.column_dimensions[column].width = width
            cell += 1
            self.current_sheet[f'{column}{self.__row}'] = field.upper()
            self.current_sheet[f'{column}{self.__row}'].fill = fill
            self.current_sheet[f'{column}{self.__row}'].font = font
            self.current_sheet[f'{column}{self.__row}'].alignment = aligment
        self.__row += 1

    def __body(self, body, fields, empty):
        '''
        :param body:
        :param fields:
        :param empty:
        :return:
        '''
        if not body:
            raise ValueError('Empty body')
        if not fields:
            raise ValueError('Empty fields')
        font = Font(name=self.__font_type)
        for data in body:
            for column in range(1, len(fields)+1):
                cel = get_column_letter(column)
                if fields[column-1] in data and data[fields[column-1]]:
                    self.current_sheet[f'{cel}{self.__row}'] = \
                        data[fields[column-1]]
                else:
                    self.current_sheet[f'{cel}{self.__row}'] = empty
                self.current_sheet[f'{cel}{self.__row}'].font = font
            self.__row += 1

    def generate_table_simple(self, fields=None, body=None, title=None,
                                header=None, footer=None, empty='Empty',
                                fill_fields='365F91', fond_fields='FFFFFF',
                                fields_alignment_h='center',
                                fields_alignment_v='center'
                                ):
            '''

            :param fields:
            :param body:
            :param title:
            :param header:
            :param footer:
            :param empty:
            :param fill_fields:
            :param fond_fields:
            :param fields_alignment_h:
            :param fields_alignment_v:
            :return: File .xlsx
            '''
            try:
                fields_length = len(fields)
                if header:
                    self.__header(title=title, fields_length=fields_length)

                self.__fields_name(fields=fields, fill_fields=fill_fields,
                                fond_fields=fond_fields,
                                fields_alignment_h=fields_alignment_h,
                                fields_alignment_v=fields_alignment_v)
                self.__body(body, fields, empty)
                if footer:
                    pass
                return {'data': self.__work_book, 'code': 200}

            except ValueError as e:
                print('Error', e)
                return {'data': f'Error de valor: {e}', 'code': 404} 